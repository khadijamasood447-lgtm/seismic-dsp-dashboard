from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows


def add_table(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def add_image(ws, image_path: Path, cell: str, width: int | None = None) -> None:
    if not image_path.exists():
        return
    img = Image(str(image_path))
    if width is not None and img.width:
        scale = width / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    ws.add_image(img, cell)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    ds_dir = root / "data" / "training" / "cyclic_triaxial_v4"
    plots_dir = ds_dir / "plots"

    metrics_csv = plots_dir / "test_metrics_summary.csv"
    if not metrics_csv.exists():
        raise SystemExit("Missing test_metrics_summary.csv; run plot scripts first")

    df_metrics = pd.read_csv(metrics_csv)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    add_table(ws_summary, df_metrics.sort_values("r2", ascending=False), start_row=1, start_col=1)

    ws_summary.column_dimensions["A"].width = 16
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 14
    ws_summary.column_dimensions["D"].width = 14

    add_image(ws_summary, plots_dir / "mae_by_target_test.png", "F2", width=700)

    ws_curves = wb.create_sheet("LearningCurves")
    add_image(ws_curves, plots_dir / "learning_curve_N_eps_0_1.png", "A1", width=800)
    add_image(ws_curves, plots_dir / "learning_curve_N_eps_1_0.png", "A30", width=800)

    ws_targets = wb.create_sheet("TargetPlots")
    targets = df_metrics["target"].tolist()
    row = 1
    for t in targets:
        scatter = plots_dir / f"scatter_{t}.png"
        resid = plots_dir / f"residuals_{t}.png"
        ws_targets.cell(row=row, column=1, value=t)
        add_image(ws_targets, scatter, f"A{row+1}", width=450)
        add_image(ws_targets, resid, f"H{row+1}", width=450)
        row += 22

    out_path = ds_dir / "accuracy_report.xlsx"
    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()

